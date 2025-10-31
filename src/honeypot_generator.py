#!/usr/bin/env python3
"""
Honeypot File Generator for ZeroPain Security System
=====================================================
Generates realistic but completely fake medical and financial files
designed to waste attacker time and cause maximum frustration.
"""

import time
import random
import json
import csv
import secrets
from datetime import datetime
from pathlib import Path
import zipfile

# Document generation
from faker import Faker
from mimesis import Person, Address, Finance, Datetime, Text
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas

# Initialize fake data generators
fake = Faker(["en_US"])
person = Person("en")
address = Address("en")
finance = Finance("en")
dt = Datetime()
text = Text("en")

# ==============================================================================
# MEDICAL DATA GENERATORS
# ==============================================================================


class MedicalDataGenerator:
    """Generates fake medical records and patient data"""

    @staticmethod
    def generate_patient_record():
        """Generate a complete patient record"""
        return {
            "patient_id": f"PT{random.randint(100000, 999999)}",
            "mrn": f"MRN{random.randint(10000000, 99999999)}",
            "name": {
                "first": person.first_name(),
                "middle": person.first_name(),
                "last": person.last_name(),
                "suffix": random.choice(["", "Jr.", "Sr.", "III", ""]),
            },
            "demographics": {
                "dob": fake.date_of_birth(minimum_age=18, maximum_age=95).isoformat(),
                "ssn": f"{random.randint(100,999)}-{random.randint(10,99)}-{random.randint(1000,9999)}",
                "gender": random.choice(["M", "F", "O"]),
                "race": random.choice(["White", "Black", "Asian", "Hispanic", "Other"]),
                "ethnicity": random.choice(["Hispanic", "Non-Hispanic"]),
                "marital_status": random.choice(
                    ["Single", "Married", "Divorced", "Widowed"]
                ),
            },
            "contact": {
                "address": address.address(),
                "city": address.city(),
                "state": address.state(),
                "zip": address.zip_code(),
                "phone": person.telephone(),
                "email": person.email(),
                "emergency_contact": {
                    "name": person.full_name(),
                    "relationship": random.choice(
                        ["Spouse", "Parent", "Sibling", "Child", "Friend"]
                    ),
                    "phone": person.telephone(),
                },
            },
            "insurance": {
                "primary": {
                    "carrier": random.choice(
                        [
                            "Blue Cross Blue Shield",
                            "Aetna",
                            "Cigna",
                            "United Healthcare",
                            "Humana",
                        ]
                    ),
                    "policy_number": f"POL{random.randint(10000000, 99999999)}",
                    "group_number": f"GRP{random.randint(1000, 9999)}",
                    "subscriber_id": f"SUB{random.randint(100000, 999999)}",
                    "copay": random.choice([10, 15, 20, 25, 30, 35, 40]),
                    "deductible": random.choice([500, 1000, 1500, 2000, 2500, 3000]),
                }
            },
            "medical_history": {
                "diagnoses": [
                    {
                        "code": f"M{random.randint(10,99)}.{random.randint(0,9)}",
                        "description": random.choice(
                            [
                                "Chronic pain syndrome",
                                "Post-laminectomy syndrome",
                                "Fibromyalgia",
                                "Complex regional pain syndrome",
                                "Neuropathic pain",
                                "Cancer pain",
                                "Osteoarthritis",
                            ]
                        ),
                        "date": fake.date_between(
                            start_date="-5y", end_date="today"
                        ).isoformat(),
                    }
                    for _ in range(random.randint(1, 5))
                ],
                "surgeries": [
                    {
                        "procedure": random.choice(
                            [
                                "Lumbar laminectomy",
                                "Spinal fusion",
                                "Total knee replacement",
                                "Hip replacement",
                                "Shoulder surgery",
                            ]
                        ),
                        "date": fake.date_between(
                            start_date="-3y", end_date="-6m"
                        ).isoformat(),
                        "surgeon": f"Dr. {person.last_name()}",
                    }
                    for _ in range(random.randint(0, 3))
                ],
                "allergies": random.sample(
                    [
                        "Penicillin",
                        "Sulfa",
                        "Morphine",
                        "Codeine",
                        "Latex",
                        "Iodine",
                        "None known",
                    ],
                    random.randint(0, 3),
                ),
            },
            "medications": {
                "current": [
                    {
                        "name": med,
                        "dose": f"{random.randint(5,100)}mg",
                        "frequency": random.choice(
                            ["QD", "BID", "TID", "QID", "Q6H", "Q8H", "Q12H"]
                        ),
                        "route": random.choice(["PO", "IV", "IM", "SC", "SL"]),
                        "prescriber": f"Dr. {person.last_name()}",
                    }
                    for med in random.sample(
                        [
                            "SR-17018",
                            "SR-14968",
                            "Oxycodone",
                            "Gabapentin",
                            "Pregabalin",
                            "Duloxetine",
                            "Amitriptyline",
                            "Tramadol",
                        ],
                        random.randint(3, 6),
                    )
                ]
            },
            "vitals": {
                "height": f"{random.randint(150,200)} cm",
                "weight": f"{random.randint(50,120)} kg",
                "bmi": round(random.uniform(18.5, 35.0), 1),
                "blood_pressure": f"{random.randint(110,180)}/{random.randint(70,110)}",
                "heart_rate": random.randint(60, 100),
                "temperature": round(random.uniform(97.0, 99.5), 1),
                "pain_score": random.randint(0, 10),
            },
            "lab_results": {
                "cbc": {
                    "wbc": round(random.uniform(4.5, 11.0), 1),
                    "rbc": round(random.uniform(4.2, 5.9), 1),
                    "hgb": round(random.uniform(12.0, 17.5), 1),
                    "hct": round(random.uniform(36.0, 50.0), 1),
                    "platelets": random.randint(150, 400),
                },
                "chemistry": {
                    "glucose": random.randint(70, 140),
                    "creatinine": round(random.uniform(0.6, 1.3), 1),
                    "bun": random.randint(7, 25),
                    "sodium": random.randint(135, 145),
                    "potassium": round(random.uniform(3.5, 5.0), 1),
                },
            },
            "clinical_trials": {
                "enrolled": random.choice([True, False]),
                "trial_id": f"NCT{random.randint(10000000, 99999999)}",
                "protocol": "ZeroPain-001",
                "randomization": random.choice(["Treatment", "Control", "Placebo"]),
                "start_date": fake.date_between(
                    start_date="-6m", end_date="today"
                ).isoformat(),
            },
        }

    @staticmethod
    def generate_clinical_trial_data(num_patients=100):
        """Generate fake clinical trial dataset"""
        data = []
        for i in range(num_patients):
            patient = MedicalDataGenerator.generate_patient_record()
            trial_data = {
                "subject_id": f"SUBJ{i:04d}",
                "patient_id": patient["patient_id"],
                "mrn": patient["mrn"],
                "name": f"{patient['name']['first']} {patient['name']['last']}",
                "dob": patient["demographics"]["dob"],
                "site": random.choice(["Site A", "Site B", "Site C", "Site D"]),
                "enrollment_date": fake.date_between(
                    start_date="-1y", end_date="today"
                ).isoformat(),
                "treatment_arm": random.choice(
                    ["SR-17018", "SR-14968", "Combination", "Placebo"]
                ),
                "baseline_pain": random.randint(6, 10),
                "week_1_pain": random.randint(3, 8),
                "week_2_pain": random.randint(2, 7),
                "week_4_pain": random.randint(1, 6),
                "week_8_pain": random.randint(0, 5),
                "week_12_pain": random.randint(0, 4),
                "adverse_events": random.choice(
                    ["None", "Mild nausea", "Constipation", "Dizziness", "Headache"]
                ),
                "serious_adverse_events": random.choice(
                    ["None"] * 95 + ["Hospitalization"] * 5
                ),
                "withdrawal": random.choice([False] * 85 + [True] * 15),
                "withdrawal_reason": random.choice(
                    [
                        "N/A",
                        "Lack of efficacy",
                        "Adverse event",
                        "Lost to follow-up",
                        "Protocol violation",
                    ]
                )
                if random.random() < 0.15
                else "N/A",
            }
            data.append(trial_data)
        return data


# ==============================================================================
# FINANCIAL DATA GENERATORS
# ==============================================================================


class FinancialDataGenerator:
    """Generates fake financial records"""

    @staticmethod
    def generate_bank_account():
        """Generate fake bank account details"""
        return {
            "account_number": f"{random.randint(1000000000, 9999999999)}",
            "routing_number": f"{random.randint(100000000, 999999999)}",
            "account_type": random.choice(
                ["Checking", "Savings", "Business", "Investment"]
            ),
            "bank_name": random.choice(
                [
                    "Chase Bank",
                    "Bank of America",
                    "Wells Fargo",
                    "Citibank",
                    "US Bank",
                    "PNC Bank",
                    "TD Bank",
                ]
            ),
            "swift_code": f"{fake.lexify('????').upper()}{fake.lexify('??').upper()}{random.randint(10,99)}",
            "iban": f"US{random.randint(10,99)}{fake.lexify('????').upper()}{random.randint(10000000000000, 99999999999999)}",
            "balance": f"${random.randint(10000, 10000000)}.{random.randint(0,99):02d}",
            "currency": "USD",
            "opened_date": fake.date_between(
                start_date="-10y", end_date="-1y"
            ).isoformat(),
            "last_transaction": fake.date_between(
                start_date="-7d", end_date="today"
            ).isoformat(),
        }

    @staticmethod
    def generate_financial_report():
        """Generate fake financial report"""
        return {
            "company": "ZeroPain Therapeutics LLC",
            "ein": f"{random.randint(10,99)}-{random.randint(1000000,9999999)}",
            "fiscal_year": 2024,
            "revenue": {
                "q1": random.randint(10000000, 50000000),
                "q2": random.randint(12000000, 55000000),
                "q3": random.randint(15000000, 60000000),
                "q4": random.randint(20000000, 70000000),
                "total": random.randint(57000000, 235000000),
            },
            "expenses": {
                "r_and_d": random.randint(20000000, 80000000),
                "clinical_trials": random.randint(30000000, 100000000),
                "manufacturing": random.randint(10000000, 40000000),
                "sales_marketing": random.randint(5000000, 20000000),
                "general_admin": random.randint(5000000, 15000000),
            },
            "assets": {
                "cash": random.randint(50000000, 200000000),
                "investments": random.randint(20000000, 100000000),
                "property": random.randint(10000000, 50000000),
                "intellectual_property": random.randint(100000000, 500000000),
            },
            "liabilities": {
                "accounts_payable": random.randint(1000000, 10000000),
                "long_term_debt": random.randint(0, 50000000),
                "deferred_revenue": random.randint(5000000, 30000000),
            },
            "bank_accounts": [
                FinancialDataGenerator.generate_bank_account()
                for _ in range(random.randint(3, 8))
            ],
            "credit_cards": [
                {
                    "number": finance.credit_card_number(),
                    "cvv": f"{random.randint(100,999)}",
                    "expiry": f"{random.randint(1,12):02d}/{random.randint(25,30)}",
                    "limit": random.randint(50000, 500000),
                    "balance": random.randint(0, 100000),
                }
                for _ in range(random.randint(2, 5))
            ],
        }


# ==============================================================================
# FILE GENERATORS
# ==============================================================================


class FileGenerator:
    """Generates various file types with fake data"""

    @staticmethod
    def generate_excel_file(filename: str, data_type: str = "medical"):
        """Generate Excel file with fake data"""
        wb = Workbook()
        ws = wb.active

        if data_type == "medical":
            # Generate patient data
            ws.title = "Patient Records"
            patients = [
                MedicalDataGenerator.generate_patient_record()
                for _ in range(random.randint(100, 1000))
            ]

            # Headers
            headers = [
                "Patient ID",
                "Name",
                "DOB",
                "SSN",
                "Diagnosis",
                "Medications",
                "Insurance",
            ]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(
                    start_color="366092", end_color="366092", fill_type="solid"
                )

            # Data
            for row, patient in enumerate(patients, 2):
                ws.cell(row=row, column=1, value=patient["patient_id"])
                ws.cell(
                    row=row,
                    column=2,
                    value=f"{patient['name']['first']} {patient['name']['last']}",
                )
                ws.cell(row=row, column=3, value=patient["demographics"]["dob"])
                ws.cell(row=row, column=4, value=patient["demographics"]["ssn"])
                ws.cell(
                    row=row,
                    column=5,
                    value=patient["medical_history"]["diagnoses"][0]["description"],
                )
                ws.cell(
                    row=row,
                    column=6,
                    value=", ".join(
                        [m["name"] for m in patient["medications"]["current"]]
                    ),
                )
                ws.cell(
                    row=row, column=7, value=patient["insurance"]["primary"]["carrier"]
                )

        elif data_type == "financial":
            # Generate financial data
            ws.title = "Financial Report"
            report = FinancialDataGenerator.generate_financial_report()

            # Add financial data
            ws.cell(row=1, column=1, value="ZeroPain Therapeutics Financial Report")
            ws.cell(row=2, column=1, value=f"Fiscal Year {report['fiscal_year']}")

            row = 4
            for category, values in report.items():
                if isinstance(values, dict):
                    ws.cell(row=row, column=1, value=category.upper())
                    row += 1
                    for key, val in values.items():
                        ws.cell(row=row, column=2, value=key)
                        ws.cell(row=row, column=3, value=val)
                        row += 1
                    row += 1

        # Add some corruption at random point
        if random.random() < 0.3:
            corruption_row = random.randint(10, ws.max_row)
            ws.cell(row=corruption_row, column=1, value="#REF!")
            ws.cell(row=corruption_row, column=2, value="#VALUE!")
            ws.cell(row=corruption_row, column=3, value="#DIV/0!")

        wb.save(filename)

    @staticmethod
    def generate_pdf_file(filename: str, data_type: str = "medical"):
        """Generate PDF with fake data"""
        c = canvas.Canvas(filename, pagesize=letter)
        width, height = letter

        if data_type == "medical":
            # Medical report
            patient = MedicalDataGenerator.generate_patient_record()

            c.setFont("Helvetica-Bold", 16)
            c.drawString(100, height - 100, "CONFIDENTIAL MEDICAL RECORD")

            c.setFont("Helvetica", 12)
            y = height - 140

            c.drawString(
                100, y, f"Patient: {patient['name']['first']} {patient['name']['last']}"
            )
            y -= 20
            c.drawString(100, y, f"MRN: {patient['mrn']}")
            y -= 20
            c.drawString(100, y, f"DOB: {patient['demographics']['dob']}")
            y -= 20
            c.drawString(100, y, f"SSN: {patient['demographics']['ssn']}")
            y -= 30

            c.setFont("Helvetica-Bold", 14)
            c.drawString(100, y, "Current Medications:")
            y -= 20

            c.setFont("Helvetica", 11)
            for med in patient["medications"]["current"]:
                c.drawString(
                    120, y, f"â€¢ {med['name']} {med['dose']} {med['frequency']}"
                )
                y -= 18

        elif data_type == "financial":
            # Financial report
            report = FinancialDataGenerator.generate_financial_report()

            c.setFont("Helvetica-Bold", 16)
            c.drawString(100, height - 100, "CONFIDENTIAL FINANCIAL REPORT")

            c.setFont("Helvetica", 12)
            y = height - 140

            c.drawString(100, y, f"Company: {report['company']}")
            y -= 20
            c.drawString(100, y, f"EIN: {report['ein']}")
            y -= 20
            c.drawString(100, y, f"Fiscal Year: {report['fiscal_year']}")
            y -= 30

            c.setFont("Helvetica-Bold", 14)
            c.drawString(100, y, "Revenue Summary:")
            y -= 20

            c.setFont("Helvetica", 11)
            for quarter, amount in report["revenue"].items():
                if quarter != "total":
                    c.drawString(120, y, f"â€¢ {quarter.upper()}: ${amount:,}")
                    y -= 18

        # Add corruption marker
        if random.random() < 0.3:
            c.setFont("Helvetica", 8)
            c.drawString(
                width / 2,
                50,
                "ERROR: PDF STRUCTURE CORRUPTED AT BYTE 0x"
                + secrets.token_hex(4).upper(),
            )

        c.save()

    @staticmethod
    def generate_corrupted_zip(filename: str):
        """Generate a ZIP file that becomes corrupted partway through"""
        # Create legitimate ZIP first
        with zipfile.ZipFile(filename, "w") as zf:
            # Add some fake files
            for i in range(random.randint(10, 50)):
                fake_filename = (
                    f"document_{i:03d}.{random.choice(['pdf', 'xlsx', 'docx', 'csv'])}"
                )
                fake_content = secrets.token_bytes(random.randint(1000, 100000))
                zf.writestr(fake_filename, fake_content)

        # Now corrupt it
        if random.random() < 0.7:
            with open(filename, "r+b") as f:
                # Corrupt the central directory
                f.seek(-22, 2)  # Seek to end of central directory
                f.write(b"CORRUPTED")


# ==============================================================================
# MAIN GENERATOR LOOP
# ==============================================================================


def main():
    """Main honeypot generation loop"""

    output_dir = Path("/output")
    output_dir.mkdir(exist_ok=True)

    print("ðŸ¯ ZeroPain Honeypot Generator Started")
    print("=" * 50)

    while True:
        try:
            # Generate medical files
            print(f"Generating medical honeypot files...")

            # Patient records Excel
            excel_file = (
                output_dir / f"patient_records_{datetime.now().strftime('%Y%m%d')}.xlsx"
            )
            FileGenerator.generate_excel_file(str(excel_file), "medical")
            print(f"  âœ“ Generated: {excel_file.name}")

            # Clinical trial data CSV
            csv_file = output_dir / f"clinical_trial_data_{secrets.token_hex(4)}.csv"
            trial_data = MedicalDataGenerator.generate_clinical_trial_data(
                random.randint(500, 2000)
            )
            with open(csv_file, "w", newline="") as f:
                if trial_data:
                    writer = csv.DictWriter(f, fieldnames=trial_data[0].keys())
                    writer.writeheader()
                    writer.writerows(trial_data)
            print(f"  âœ“ Generated: {csv_file.name}")

            # Medical reports PDF
            pdf_file = output_dir / f"medical_report_{secrets.token_hex(4)}.pdf"
            FileGenerator.generate_pdf_file(str(pdf_file), "medical")
            print(f"  âœ“ Generated: {pdf_file.name}")

            # Generate financial files
            print(f"Generating financial honeypot files...")

            # Financial report Excel
            excel_file = (
                output_dir / f"financial_report_Q{random.randint(1,4)}_2024.xlsx"
            )
            FileGenerator.generate_excel_file(str(excel_file), "financial")
            print(f"  âœ“ Generated: {excel_file.name}")

            # Bank statements PDF
            pdf_file = (
                output_dir / f"bank_statement_{datetime.now().strftime('%Y%m')}.pdf"
            )
            FileGenerator.generate_pdf_file(str(pdf_file), "financial")
            print(f"  âœ“ Generated: {pdf_file.name}")

            # Generate corrupted archives
            print(f"Generating corrupted archives...")

            # Corrupted patient data ZIP
            zip_file = (
                output_dir / f"patient_database_backup_{secrets.token_hex(4)}.zip"
            )
            FileGenerator.generate_corrupted_zip(str(zip_file))
            print(f"  âœ“ Generated: {zip_file.name} (corrupted)")

            # JSON dumps with progressive corruption
            json_file = output_dir / f"api_dump_{secrets.token_hex(4)}.json"
            data = {
                "patients": [
                    MedicalDataGenerator.generate_patient_record()
                    for _ in range(random.randint(100, 500))
                ],
                "timestamp": datetime.now().isoformat(),
            }

            # Start writing valid JSON then corrupt it
            with open(json_file, "w") as f:
                json_str = json.dumps(data, indent=2)
                # Write most of it correctly
                f.write(json_str[: int(len(json_str) * 0.7)])
                # Then add corruption
                f.write("\n\nERROR: DATABASE CONNECTION LOST\n")
                f.write("null" * 1000)
                f.write("\n0xDEADBEEF" * 100)
            print(f"  âœ“ Generated: {json_file.name} (corrupted)")

            # Clean up old files (older than 24 hours)
            for old_file in output_dir.glob("*"):
                if old_file.is_file():
                    file_age = time.time() - old_file.stat().st_mtime
                    if file_age > 86400:  # 24 hours
                        old_file.unlink()
                        print(f"  ðŸ—‘ Cleaned up: {old_file.name}")

            print(f"\nâœ… Honeypot generation complete. Sleeping for 1 hour...\n")
            time.sleep(3600)  # Generate new files every hour

        except Exception as e:
            print(f"âŒ Error during generation: {e}")
            time.sleep(60)  # Wait a minute before retrying


if __name__ == "__main__":
    main()
